from flask import Flask
import pymysql
import openpyxl

app = Flask(__name__)
app.config.from_object('testapp.config')

conn = pymysql.connect(host='localhost',
                       user='t4',
                       password='t4_password',
                       database='the1',
                       cursorclass=pymysql.cursors.DictCursor)
cursor = conn.cursor()


sql = "CREATE TABLE IF NOT EXISTS asp_men (id INT UNIQUE, player VARCHAR(255), z1 INT, z2 INT, z3 INT, z4 INT, t1 INT, t2 INT, t3 INT, t4 INT, total numeric)"
cursor.execute(sql)
conn.commit()
sql = "CREATE TABLE IF NOT EXISTS asp_wmn (id INT UNIQUE, player VARCHAR(255), z1 INT, z2 INT, z3 INT, z4 INT, t1 INT, t2 INT, t3 INT, t4 INT, total numeric)"
cursor.execute(sql)
conn.commit()
sql = "CREATE TABLE IF NOT EXISTS fin_men (id INT UNIQUE, player VARCHAR(255), z1 INT, z2 INT, z3 INT, z4 INT, t1 INT, t2 INT, t3 INT, t4 INT, total numeric)"
cursor.execute(sql)
conn.commit()
sql = "CREATE TABLE IF NOT EXISTS fin_wmn (id INT UNIQUE, player VARCHAR(255), z1 INT, z2 INT, z3 INT, z4 INT, t1 INT, t2 INT, t3 INT, t4 INT, total numeric)"
cursor.execute(sql)
conn.commit()

print("--------------------------------------")
print("[DB] > 初期化完了")
print("--------------------------------------")

wb = openpyxl.load_workbook("static.xlsx")

ws = wb["asp_men"]
for i in range(1,100):
    temp = (ws.cell(i,1).value, ws.cell(i,2).value)
    sql = "INSERT INTO asp_men(id, player) VALUES (%s, %s) ON DUPLICATE KEY UPDATE player = VALUES(player)"
    cursor.execute(sql, temp)
    conn.commit()

ws = wb["asp_wmn"]
for i in range(1,100):
    temp = (ws.cell(i,1).value, ws.cell(i,2).value)
    sql = "INSERT INTO asp_wmn(id, player) VALUES (%s, %s) ON DUPLICATE KEY UPDATE player = VALUES(player)"
    cursor.execute(sql, temp)
    conn.commit()

ws = wb["fin_men"]
for i in range(1,100):
    temp = (ws.cell(i,1).value, ws.cell(i,2).value)
    sql = "INSERT INTO fin_men(id, player) VALUES (%s, %s) ON DUPLICATE KEY UPDATE player = VALUES(player)"
    cursor.execute(sql, temp)
    conn.commit()

ws = wb["fin_wmn"]
for i in range(1,100):
    temp = (ws.cell(i,1).value, ws.cell(i,2).value)
    sql = "INSERT INTO fin_wmn(id, player) VALUES (%s, %s) ON DUPLICATE KEY UPDATE player = VALUES(player)"
    cursor.execute(sql, temp)
    conn.commit()

print("--------------------------------------")

import testapp.views