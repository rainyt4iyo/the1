import openpyxl
import os
import pymysql


def init():
    conn = pymysql.connect(host='localhost',
                        user='root',
                        password='t4_password',
                        database='myDB',
                        cursorclass=pymysql.cursors.DictCursor)

    cursor = conn.cursor()

    sql = "DROP TABLE if exists fun_player"
    cursor.execute(sql)
    sql = "DROP TABLE if exists mid_player"
    cursor.execute(sql)
    sql = "DROP TABLE if exists open_player"
    cursor.execute(sql)
    sql = "DROP TABLE if exists fun_team"
    cursor.execute(sql)
    sql = "DROP TABLE if exists mid_team"
    cursor.execute(sql)
    sql = "DROP TABLE if exists open_team"
    cursor.execute(sql)

    sql = "CREATE TABLE IF NOT EXISTS fun_player (id INT UNIQUE, player VARCHAR(255), team VARCHAR(255))"
    cursor.execute(sql)
    conn.commit()
    sql = "CREATE TABLE IF NOT EXISTS mid_player (id INT UNIQUE, player VARCHAR(255), team VARCHAR(255))"
    cursor.execute(sql)
    conn.commit()
    sql = "CREATE TABLE IF NOT EXISTS open_player (id INT UNIQUE, player VARCHAR(255), team VARCHAR(255))"
    cursor.execute(sql)
    conn.commit()

    print("DB > player テーブル作成完了")

def excel2sql():

    conn = pymysql.connect(host='localhost',
                        user='root',
                        password='t4_password',
                        database='myDB',
                        cursorclass=pymysql.cursors.DictCursor)
    cursor = conn.cursor()

    wb = openpyxl.load_workbook("player.xlsx")

    ws = wb["FUN"]
    for i in range(1,145):
        if ws.cell(i,1).value != "None" and ws.cell(i,2).value != "None" and ws.cell(i,3).value != "None":
            temp = (ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
            sql = "INSERT INTO fun_player (id, player, team ) VALUES (%s, %s, %s)"
            cursor.execute(sql, temp)
            conn.commit()

    ws = wb["MIDDLE"]
    for i in range(1,145):
        if ws.cell(i,1).value != "None" and ws.cell(i,2).value != "None" and ws.cell(i,3).value != "None":
            temp = (ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
            sql = "INSERT INTO mid_player (id, player, team ) VALUES (%s, %s, %s)"
            cursor.execute(sql, temp)
            conn.commit()

    ws = wb["OPEN"]
    for i in range(1,145):
        if ws.cell(i,1).value != "None" and ws.cell(i,2).value != "None" and ws.cell(i,3).value != "None":
            temp = (ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
            sql = "INSERT INTO open_player (id, player, team ) VALUES (%s, %s, %s)"
            cursor.execute(sql, temp)
            conn.commit()
    conn.close()


def main():
    init()
    excel2sql()

if __name__ == '__main__':
    main()
    
